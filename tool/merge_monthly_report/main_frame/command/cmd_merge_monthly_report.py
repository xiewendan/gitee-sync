# -*- coding: utf-8 -*-

# __author__ = xiaobao
# __date__ = 2020/4/18 8:58 下午

# desc: 

import datetime
import os
from copy import copy

import openpyxl

import common.my_exception as my_exception
import common.my_path as my_path
import main_frame.cmd_base as cmd_base
from openpyxl.worksheet.datavalidation import DataValidation


def GetLastMonth():
    nMonth = datetime.datetime.now().month
    nLastMonth = nMonth - 1

    if nLastMonth == 0:
        nLastMonth = 12

    return nLastMonth


def CopyCell(SrcCellObj, DestCellObj):
    DestCellObj.value = SrcCellObj.value
    DestCellObj.data_type = SrcCellObj.data_type
    DestCellObj.fill = copy(SrcCellObj.fill)
    if SrcCellObj.has_style:
        # DestCellObj._style = copy(SrcCellObj._style)
        DestCellObj.font = copy(SrcCellObj.font)
        DestCellObj.border = copy(SrcCellObj.border)
        DestCellObj.fill = copy(SrcCellObj.fill)
        DestCellObj.number_format = copy(SrcCellObj.number_format)
        DestCellObj.protection = copy(SrcCellObj.protection)
        DestCellObj.alignment = copy(SrcCellObj.alignment)

    if SrcCellObj.hyperlink:
        DestCellObj._hyperlink = copy(SrcCellObj.hyperlink)

    if SrcCellObj.comment:
        DestCellObj.comment = copy(SrcCellObj.comment)


class SheetName:
    eSummary = "汇总表"
    eTemplate = "模板"


class CmdMergeMonthlyReport(cmd_base.CmdBase):
    def __init__(self):
        self.m_AppObj = None

    @staticmethod
    def GetName():
        return "merge_monthly_report"

    def Init(self, AppObj):
        self.m_AppObj = AppObj

    def Do(self):
        self.m_AppObj.Info("Start MergeMonthlyReport")

        szCWD = self.m_AppObj.ConfigLoader.CWD

        szMrTemplatePath = szCWD + "/" + self.m_AppObj.CLM.GetArg(1)
        szSrcDir = szCWD + "/" + self.m_AppObj.CLM.GetArg(2)
        szDestDir = szCWD + "/" + self.m_AppObj.CLM.GetArg(3)
        self.m_AppObj.Info(
            "\ntemplate path:{0}\nsrc path:{1}\ndest path:{2}\n".format(szMrTemplatePath, szSrcDir, szDestDir))

        szMonth = str(datetime.datetime.now().month)
        szFileName = my_path.FileNameWithExt(szMrTemplatePath)
        szDestPath = "{0}/{1}".format(szDestDir, szFileName.replace("xx", szMonth))
        self.m_AppObj.Info("dest path:{0}".format(szDestPath))

        MrWorkbookObj = MrWorkbook(self.m_AppObj, szMrTemplatePath, szSrcDir)
        MrWorkbookObj.Handle()
        MrWorkbookObj.Save(szDestPath)
        MrWorkbookObj.Close()

        self.m_AppObj.Info("End MergeMonthlyReport")


# MonthlyReporWorkbook
class MrWorkbook:
    def __init__(self, AppObj, szTemplatePath, szSrcDir):
        self.m_AppObj = AppObj
        self.m_WorkbookObj = openpyxl.load_workbook(szTemplatePath)
        self.m_szSrcDir = szSrcDir
        self.m_listMemberName = []

    def CheckData(self):
        listSheetName = self.m_WorkbookObj.sheetnames
        self.m_AppObj.Info("{0}, {1}".format(SheetName.eSummary, SheetName.eTemplate))

        assert SheetName.eSummary in listSheetName and SheetName.eTemplate in listSheetName, \
            "tempalte excel missing sheet {0} or {1},  sheet name list: {2}".format(
                SheetName.eSummary, SheetName.eTemplate, ",".join(listSheetName))

        self.m_listMemberName = []
        SummarySheetObj = self.m_WorkbookObj[SheetName.eSummary]
        nMaxColumn = SummarySheetObj.max_column
        for nColumnIndex in range(2, nMaxColumn + 1):
            szColumnChar = chr(96 + nColumnIndex)
            szCellPos1 = "{0}1".format(szColumnChar)
            szMemberName = SummarySheetObj[szCellPos1].value
            if szMemberName is not None:
                self.m_listMemberName.append(szMemberName)

        listMissingMemberExcel = []
        for szMemberName in self.m_listMemberName:
            szMemberFilePath = "{0}/{1}.xlsx".format(self.m_szSrcDir, szMemberName)
            if not os.path.exists(szMemberFilePath):
                listMissingMemberExcel.append(szMemberName)
                self.m_AppObj.Error("missing excel:{0}".format(szMemberFilePath))

        if len(listMissingMemberExcel) > 0:
            self.m_AppObj.Error("member missing excel:{0}".format(",".join(listMissingMemberExcel)))
            return False

        return True

    def Handle(self):
        self.m_AppObj.Debug("begin workbook handle")

        if not self.CheckData():
            raise my_exception.MyException("missing member excel")

        TemplateSheetObj = self.m_WorkbookObj[SheetName.eTemplate]
        for szMemberName in self.m_listMemberName:
            szMemberFilePath = "{0}/{1}.xlsx".format(self.m_szSrcDir, szMemberName)
            MemberWorkbookObj = openpyxl.load_workbook(szMemberFilePath)
            szFirstSheetName = MemberWorkbookObj.sheetnames[0]
            SrcMemberSheetObj = MemberWorkbookObj[szFirstSheetName]

            MemberSheetObj = self.m_WorkbookObj.copy_worksheet(TemplateSheetObj)
            MemberSheetObj.title = szMemberName

            TmSheetObj = TmSheet(self.m_AppObj, MemberSheetObj, SrcMemberSheetObj)
            TmSheetObj.Handle()

        SheetObjInBook = self.m_WorkbookObj[SheetName.eSummary]
        SummarySheetObj = SummarySheet(self.m_AppObj, SheetObjInBook, self.m_listMemberName)
        SummarySheetObj.Handle()

        self.m_AppObj.Debug("end workbook handle")

    def Save(self, szPath):
        self.m_WorkbookObj.save(szPath)

    def Close(self):
        self.m_WorkbookObj.close()
        self.m_WorkbookObj = None


# TeamMemberSheet：每个成员表
class TmSheet:
    def __init__(self, AppObj, SheetObj, SrcSheetObj):
        self.m_AppObj = AppObj
        self.m_SheetObj = SheetObj
        self.m_SrcSheetObj = SrcSheetObj
        self.m_nMaxRow = 0

    def Handle(self):
        self.m_AppObj.Debug("begin tmsheet handle:{0}".format(self.m_SheetObj.title))

        # 数据修改
        self.UpdateData()

        # 格式修改
        self.UpdateFormat()

        # 数据检查
        self.CheckDataValidate()

        self.m_AppObj.Debug("end tmsheet handle:{0}".format(self.m_SheetObj.title))

    def UpdateData(self):
        self.m_AppObj.Debug("update data")

        nMaxRow = self.m_SrcSheetObj.max_row
        # 时间
        self.m_SheetObj["A2"].value = "{0}月".format(GetLastMonth())

        # 工作内容
        # 天数
        for nRowIndex in range(2, nMaxRow + 1):
            szCellPos = "B{0}".format(nRowIndex)
            CopyCell(self.m_SrcSheetObj[szCellPos], self.m_SheetObj[szCellPos])

            szCellPos = "C{0}".format(nRowIndex)
            CopyCell(self.m_SrcSheetObj[szCellPos], self.m_SheetObj[szCellPos])

            if self.m_SheetObj[szCellPos].value is not None:
                self.m_nMaxRow = nRowIndex

        self.m_AppObj.Info("max row:{0}".format(self.m_nMaxRow))

        # 重要程度
        # 代码质量
        # 实现难度
        # 交付效率
        listCharacter = ["D", "E", "F", "G"]
        for nRowIndex in range(3, self.m_nMaxRow + 1):
            for szChar in listCharacter:
                szDefaultPos = "{0}2".format(szChar)
                szCellPos = "{0}{1}".format(szChar, nRowIndex)
                CopyCell(self.m_SheetObj[szDefaultPos], self.m_SheetObj[szCellPos])

        # 评分：公式
        szScoreFormat = "=C{0} * IF(D{0} =\"核心\",1.3,IF(D{0}=\"基本\",1.1,IF(D{0}=\"次要\",0.9,IF(D{0}=\"周边\",0.7,IF(D{0}=\"改bug\",0.5,IF(D{0}=\"自学\",0.2,IF(D{0}=\"无关\",0,0))))))) \
            * IF(E{0} =\"超水准\",1.3,IF(E{0}=\"基本达标\",1,IF(E{0}=\"少量问题\",0.8,IF(E{0}=\"引发事故\",0.6,IF(E{0}=\"\",0))))) \
            * IF(F{0} =\"噩梦\",1.5,IF(F{0}=\"困难\",1.3,IF(F{0}=\"普通\",1,IF(F{0}=\"简单\",0.8,IF(F{0}=\"小白\",0.6,0)))) \
            * IF(G{0} =\"超前\",1.2,IF(G{0}=\"按时\",1,IF(G{0}=\"稍晚\",0.8,IF(G{0}=\"延期\",0.6,IF(G{0}=\"中止\",0.5,0))))))"

        for nRowIndex in range(2, self.m_nMaxRow + 1):
            szCellPos = "H{0}".format(nRowIndex)
            self.m_SheetObj[szCellPos].value = szScoreFormat.format(nRowIndex)

        # 微调

        # 总计
        self.m_SheetObj["J2"] = "=SUM(H2:H{0}, I2)".format(nMaxRow)

    def UpdateFormat(self):
        self.m_AppObj.Debug("update format")

        self.m_SheetObj["A2"].value = "{0}月".format(GetLastMonth())

        nMaxRow = self.m_nMaxRow
        self.m_AppObj.Debug("max row:{0}".format(nMaxRow))

        # 时间合并单元格
        self.m_SheetObj.merge_cells('A2:A{0}'.format(nMaxRow))

        # 微调合并单元格
        self.m_SheetObj.merge_cells('I2:I{0}'.format(nMaxRow))

        # 总计合并单元格
        self.m_SheetObj.merge_cells('J2:J{0}'.format(nMaxRow))

        # 更新数据验证
        dictDataValidation = {
            "D": '"核心,基本,次要,周边,改bug,自学,无关"',
            "E": '"超水准,基本达标,少量问题,引发事故"',
            "F": '"噩梦,困难,普通,简单,小白"',
            "G": '"超前,按时,稍晚,延期,中止"'
        }
        for szColName, szData in dictDataValidation.items():
            DataValidationObj = DataValidation(type="list", formula1=szData, allow_blank=True)
            DataValidationObj.add('{0}1:{0}1048576'.format(szColName))
            self.m_SheetObj.add_data_validation(DataValidationObj)

    def CheckDataValidate(self):
        self.m_AppObj.Debug("update data validate")

        # 天数检查
        nSum = 0
        nMaxRow = self.m_nMaxRow
        for nRowIndex in range(2, nMaxRow + 1):
            nSum += int(self.m_SheetObj["C{0}".format(nRowIndex)].value)

        if nSum > 22:
            self.m_AppObj.Error("天数总和异常:{0}, {1}".format(self.m_SheetObj.title, nSum))


# 汇总表
class SummarySheet:
    def __init__(self, AppObj, SheetObj, listMemberName):
        self.m_AppObj = AppObj
        self.m_SheetObj = SheetObj
        self.m_listMemberName = listMemberName

    def Handle(self):
        self.m_AppObj.Debug("Begin Handle")

        nMaxColumn = len(self.m_listMemberName) + 1
        for nColumnIndex in range(2, nMaxColumn + 1):
            szColumnChar = chr(96 + nColumnIndex)
            szCellPos1 = "{0}1".format(szColumnChar)
            szCellPos2 = "{0}2".format(szColumnChar)
            self.m_SheetObj[szCellPos2].value = "={0}!J2".format(self.m_SheetObj[szCellPos1].value)


# 单元格填充自动拓展算法：
# 1、匹配特殊序列
# 2、增长规律
