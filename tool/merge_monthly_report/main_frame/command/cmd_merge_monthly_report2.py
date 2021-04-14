# -*- coding: utf-8 -*-

# __author__ = xiaobao
# __date__ = 2020/4/18 8:58 下午

# desc: 版本2，
# 1、按照目录结构规范，存放个人的工作内容（半月），如
#   src
#       2020-12-01：是2020年12月1号到12月15号的工作内容
#       2020-12-16：是2020年12月16号到12月31号的工作内容
#       2021-01-01：是2021年01月01号到01月15号的工作内容
#       2021-01-16：是2021年01月16号到01月31号的工作内容
#       ...
#   dest
#       2020-12月月初贡献评分.xlsx
#       2020-12月月中贡献评分.xlsx
#       2021-01月月初贡献评分.xlsx
#       2021-01月月中贡献评分.xlsx
#       ...
#   template
#       xx月贡献评分-半月.xlsx
#       xx月贡献评分-月初.xlsx
#       xx月贡献评分-月中.xlsx
#
# 2、根据输入的SpecifiedDate（如果为0，则取执行脚本的当前日期），计算得到一下几个时间
#
#       工作内容-文件夹1   \  半月模板(输入)    /  工作内容-合并文件1  \    输出模板（根据当前时间计算）
#                         >===============<                       >============================= 输出文件（月初报告或月中报告）
#       工作内容-文件夹2   /                  \  工作内容-合并文件2  /      1号到10号为月初模板
#                                                                       10号到20号未月中模板
#                                                                       其余时间报错
#
#       如上图，计算一下几个路径
#       工作内容-文件夹1：data\src\日期1
#       工作内容-文件夹2：data\src\日期1
#       工作内容-合并文件1：data\src\日期1\日期1.xlsx
#       工作内容-合并文件2：data\src\日期2\日期2.xlsx
#       输出模板：
#       输出文件：data\dest目录下
#

import datetime
import logging
import os
from copy import copy

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

import common.my_exception as my_exception
import main_frame.cmd_base as cmd_base

g_ErrorLog = []


class ESheetName:
    eSummary = "汇总表"
    eTemplate = "模板"


class EColName:
    eWork = "工作内容"
    eDay = "天数"


def _GetLastMonth():
    nMonth = datetime.datetime.now().month
    nLastMonth = nMonth - 1

    if nLastMonth == 0:
        nLastMonth = 12

    return nLastMonth


def _GetCurYearMonth():
    CurDataTime = datetime.datetime.now()
    nYear = CurDataTime.year
    nMonth = CurDataTime.month
    return nYear, nMonth


def _IsCellValid(SheetCellObj):
    return SheetCellObj.value is not None and SheetCellObj.value != ""


# 以天数最大行数为准
def _GetMaxRow(SheetObj):
    nMaxRow = SheetObj.max_row
    nMaxCol = SheetObj.max_column
    szDayColChar = None
    for nColIndex in range(1, nMaxCol+1):
        szColumnChar = chr(96 + nColIndex)
        szCellPos = "{0}1".format(szColumnChar)
        if SheetObj[szCellPos].value == EColName.eDay:
            szDayColChar = szColumnChar
            break

    if szDayColChar is None:
        raise my_exception.MyException("SheetObj has not 天数: %s" % SheetObj.title)

    nValidMaxRow = 0
    for nRowIndex in range(1, nMaxRow+1):
        szCellPos = "{0}{1}".format(szDayColChar, nRowIndex)
        if _IsCellValid(SheetObj[szCellPos]):
            nValidMaxRow = nRowIndex

    return nValidMaxRow


def _CopyCell(SrcCellObj, DestCellObj):
    DestCellObj.value = SrcCellObj.value
    DestCellObj.data_type = SrcCellObj.data_type
    DestCellObj.fill = copy(SrcCellObj.fill)
    if SrcCellObj.has_style:
        # DestCellObj._style = copy(SrcCellObj._style)
        DestCellObj.font = copy(SrcCellObj.font)
        DestCellObj.border = copy(SrcCellObj.border)
        DestCellObj.fill = copy(SrcCellObj.fill)
        DestCellObj.number_format = copy(SrcCellObj.number_format)
        DestCellObj.protection = copy(SrcCellObj.protection)
        DestCellObj.alignment = copy(SrcCellObj.alignment)

    if SrcCellObj.hyperlink:
        DestCellObj._hyperlink = copy(SrcCellObj.hyperlink)

    if SrcCellObj.comment:
        DestCellObj.comment = copy(SrcCellObj.comment)


class InputData(object):
    def __init__(self):
        self.m_listSrcFDir = []
        self.m_listMergeFileFPath = []
        self.m_szMergeTemplateFPath = None
        self.m_szOutputTemplateFPath = None
        self.m_szOutputFileFPath = None

    @property
    def SrcFDir(self):
        return self.m_listSrcFDir

    @property
    def MergeFileFPath(self):
        return self.m_listMergeFileFPath

    @property
    def MergeTemplateFPath(self):
        return self.m_szMergeTemplateFPath

    @property
    def OutTemplateFPath(self):
        return self.m_szOutputTemplateFPath

    @property
    def OutputFileFPath(self):
        return self.m_szOutputFileFPath

    @staticmethod
    def _GetLastMonth(nYear, nMonth, nDay):
        assert nYear >= 0
        assert 1 <= nMonth <= 12
        assert 1 <= nDay

        DateTimeObj = datetime.datetime(nYear, nMonth, nDay)
        LastMonthDateTimeObj = DateTimeObj - datetime.timedelta(days=(nDay + 1))

        nLastMonthYear = LastMonthDateTimeObj.year
        nLastMonthMonth = LastMonthDateTimeObj.month

        return nLastMonthYear, nLastMonthMonth

    def Init(self, szDate, szDataPath):
        logging.getLogger("myLog").debug("date:%s, data path:%s" % (szDate, szDataPath))

        if szDate == "0":
            szDate = datetime.datetime.now().strftime("%Y-%m-%d")
        try:
            szYear, szMonth, szDay = szDate.split("-")
            nYear, nMonth, nDay = int(szYear), int(szMonth), int(szDay)
        except BaseException:
            raise my_exception.MyException("szDate format error，like：yyyy-mm-dd.  szDate: %s".format(szDate))

        nLastMonthYear, nLastMonthMonth = self._GetLastMonth(nYear, nMonth, nDay)

        if 1 <= nDay <= 10:
            self.m_listSrcFDir = [
                "%s/src/%04d-%02d-01" % (szDataPath, nLastMonthYear, nLastMonthMonth),
                "%s/src/%04d-%02d-16" % (szDataPath, nLastMonthYear, nLastMonthMonth),
            ]
            self.m_listMergeFileFPath = [
                "%s/src/%04d-%02d-01/%04d-%02d-01.xlsx" % (
                    szDataPath, nLastMonthYear, nLastMonthMonth, nLastMonthYear, nLastMonthMonth),
                "%s/src/%04d-%02d-16/%04d-%02d-16.xlsx" % (
                    szDataPath, nLastMonthYear, nLastMonthMonth, nLastMonthYear, nLastMonthMonth),
            ]
            self.m_szOutputTemplateFPath = "%s/template/xx月贡献评分-月初.xlsx" % szDataPath
            self.m_szOutputFileFPath = "%s/dest/%04d-%02d月贡献评分-月初.xlsx" % (szDataPath, nLastMonthYear, nLastMonthMonth)
        elif 11 <= nDay <= 20:
            self.m_listSrcFDir = [
                "%s/src/%04d-%02d-16" % (szDataPath, nLastMonthYear, nLastMonthMonth),
                "%s/src/%04d-%02d-01" % (szDataPath, nYear, nMonth),
            ]
            self.m_listMergeFileFPath = [
                "%s/src/%04d-%02d-16/%04d-%02d-16.xlsx" % (
                    szDataPath, nLastMonthYear, nLastMonthMonth, nLastMonthYear, nLastMonthMonth),
                "%s/src/%04d-%02d-01/%04d-%02d-01.xlsx" % (szDataPath, nYear, nMonth, nYear, nMonth),
            ]
            self.m_szOutputTemplateFPath = "%s/template/xx月贡献评分-月中.xlsx" % szDataPath
            self.m_szOutputFileFPath = "%s/dest/%04d-%02d月贡献评分-月中.xlsx" % (szDataPath, nYear, nMonth)
        else:
            my_exception.MyException("nDay error. 1<=nday<=20")

        self.m_szMergeTemplateFPath = "%s/template/xx月贡献评分-半月.xlsx" % szDataPath

        logging.getLogger("myLog").debug("SrcFDir:\n" + "\n".join(self.m_listSrcFDir))
        logging.getLogger("myLog").debug("MergeFileFPath:\n" + "\n".join(self.m_listMergeFileFPath))
        logging.getLogger("myLog").debug("MergeTemplateFPath:" + self.m_szMergeTemplateFPath)
        logging.getLogger("myLog").debug("OutputTemplateFPath:" + self.m_szOutputTemplateFPath)
        logging.getLogger("myLog").debug("OutputFileFPath:" + self.m_szOutputFileFPath)

    def Check(self):
        """
        1、所有文件目录存在
        2、template文件里面包含汇总表和模板表
        3、template汇总表中，所有成员名对应的成员表都存在
        4、成员表的格式都正确
        """
        # 1 目录存在
        for szSrcFDir in self.m_listSrcFDir:
            if not os.path.exists(szSrcFDir):
                g_ErrorLog.append("dir not exist:{0}".format(szSrcFDir))

        if not os.path.exists(self.m_szMergeTemplateFPath):
            g_ErrorLog.append("merge template not exist:{0}".format(self.m_szMergeTemplateFPath))

        if not os.path.exists(self.m_szOutputTemplateFPath):
            g_ErrorLog.append("output template not exist:{0}".format(self.m_szMergeTemplateFPath))

        if len(g_ErrorLog) > 0:
            logging.getLogger("myLog").error("报错信息:%s\n", "\n".join(g_ErrorLog))
            return False

        # 2 template
        listTemplate = [self.m_szMergeTemplateFPath, self.m_szOutputTemplateFPath]
        for szTemplate in listTemplate:
            if not self._CheckTemplate(szTemplate):
                return False

        return True

    def _CheckTemplate(self, szTemplate):
        logging.getLogger("myLog").debug("template:%s", szTemplate)
        TemplateWorkbookObj = openpyxl.load_workbook(szTemplate)

        listSheetName = TemplateWorkbookObj.sheetnames
        logging.getLogger("myLog").info("{0}, {1}".format(ESheetName.eSummary, ESheetName.eTemplate))

        assert ESheetName.eSummary in listSheetName and ESheetName.eTemplate in listSheetName, \
            "tempalte excel missing sheet {0} or {1},  sheet name list: {2}".format(
                ESheetName.eSummary, ESheetName.eTemplate, ",".join(listSheetName))

        # 成员名
        listMemberName = []
        SummarySheetObj = TemplateWorkbookObj[ESheetName.eSummary]
        nMaxColumn = SummarySheetObj.max_column
        for nColumnIndex in range(2, nMaxColumn + 1):
            szColumnChar = chr(96 + nColumnIndex)
            szCellPos1 = "{0}1".format(szColumnChar)
            szMemberName = SummarySheetObj[szCellPos1].value
            if szMemberName is not None:
                listMemberName.append(szMemberName)

        # 表是否存在，且检查表的格式
        listMissingPath = []
        listFormatErrorPath = []
        for szFDir in self.m_listSrcFDir:
            if not os.path.exists(szFDir):
                listMissingPath.append(szFDir)
                continue

            for szMemberName in listMemberName:
                szMemberFileFPath = os.path.join(szFDir, szMemberName + ".xlsx")
                if not os.path.exists(szMemberFileFPath):
                    listMissingPath.append(szMemberFileFPath)
                    continue

                if not self._CheckMember(szMemberFileFPath):
                    listFormatErrorPath.append(szMemberFileFPath)

        if len(listMissingPath) > 0:
            logging.getLogger("myLog").error("file missing excel:{0}".format(",".join(listMissingPath)))
            return False

        if len(listFormatErrorPath) > 0:
            return False

        return True

    @staticmethod
    def _CheckMember(szMemberFileFPath):
        logging.getLogger("myLog").debug("check member:%s", szMemberFileFPath)
        MemberWorkbookObj = openpyxl.load_workbook(szMemberFileFPath)

        szFirstSheetName = MemberWorkbookObj.sheetnames[0]
        MemberSheetObj = MemberWorkbookObj[szFirstSheetName]

        # 确认列名ok
        if MemberSheetObj["B1"].value != "工作内容":
            logging.getLogger("myLog").error("%s[B1] is not 工作内容", szMemberFileFPath)
            return False

        if MemberSheetObj["C1"].value != "天数":
            logging.getLogger("myLog").error("%s[C1] is not 天数", szMemberFileFPath)
            return False

        # 确认行列个数正确
        nMaxRow = _GetMaxRow(MemberSheetObj)

        nBRow = 0
        nCRow = 0
        for nIndex in range(2, nMaxRow+1):
            szCellPos = "B{0}".format(nIndex)
            if _IsCellValid(MemberSheetObj[szCellPos]):
                nBRow = nIndex
            szCellPos = "C{0}".format(nIndex)
            if _IsCellValid(MemberSheetObj[szCellPos]):
                nCRow = nIndex

        if nBRow != nCRow:
            logging.getLogger("myLog").error("行数不一致:%s, BRow:%d, CRow:%d", szMemberFileFPath, nBRow, nCRow)
            logging.getLogger("myLog").error("B[%d] value:%s", nBRow, str(MemberSheetObj["B{0}".format(nBRow)].value))
            logging.getLogger("myLog").error("C[%d] value:%s", nCRow, str(MemberSheetObj["C{0}".format(nCRow)].value))
            return False

        if nBRow != nMaxRow or nCRow != nMaxRow:
            logging.getLogger("myLog").error(
                "最大行数有问题:%s, MaxRow:%d, BRow:%d, CRow:%d", szMemberFileFPath, nMaxRow, nBRow, nCRow)
            return False

        # 确认天数这一列没有合并单元格
        for nIndex in range(1, nMaxRow+1):
            szCellPos = "C{0}".format(nIndex)
            if szCellPos in MemberSheetObj.merged_cells:
                logging.getLogger("myLog").error("合并单元格问题:%s, C[%d]是合并单元格", szMemberFileFPath, nCRow)
                return False

        return True


class CmdMergeMonthlyReport2(cmd_base.CmdBase):
    """
    参数1：命令名，即：merge_monthly_report
    参数2：数据源目录：data目录，目录结构参见上面
    参数3：日期：0表示采用当天日期
    参数4：有效天数：一个月有效天数根据当月工作天数，外部输入
    """

    def __init__(self):
        super().__init__()
        self.m_InputDataObj = None

    @staticmethod
    def GetName():
        return "merge_monthly_report2"

    def Init(self, AppObj):
        self.m_AppObj = AppObj

    def Do(self):
        self.m_AppObj.Info("Start MergeMonthlyReport")

        szBaseDataPath = self.m_AppObj.CLM.GetArg(1)
        logging.getLogger("myLog").info("BaseDataPath:%s", szBaseDataPath)

        szDate = self.m_AppObj.CLM.GetArg(2)
        logging.getLogger("myLog").info("Date:%s", szDate)

        nValidDay = int(self.m_AppObj.CLM.GetArg(3))
        logging.getLogger("myLog").info("ValidDay:%d", nValidDay)

        # 动态计算输入信息
        self.m_InputDataObj = InputData()
        self.m_InputDataObj.Init(szDate, szBaseDataPath)
        if not self.m_InputDataObj.Check():
            logging.getLogger("myLog").error("Input data error")
            return

        # 成员绩效转成半月绩效汇总表
        listSrcFDir = self.m_InputDataObj.SrcFDir
        listMergeFileFPath = self.m_InputDataObj.MergeFileFPath
        for nIndex, szSrcFDir in enumerate(listSrcFDir):
            szMergeFileFPath = listMergeFileFPath[nIndex]
            if os.path.exists(szMergeFileFPath):
                logging.getLogger("myLog").warning("half monthly report exist:%s", szMergeFileFPath)
                continue

            self.DoMergeMembersToHalfMonthly(
                self.m_InputDataObj.m_szMergeTemplateFPath,
                szSrcFDir,
                szMergeFileFPath,
                nValidDay
            )

        # 将多张半月绩效汇总表合并成月报
        self.DoMergeHalfMonthlyToMonthly(
            self.m_InputDataObj.OutTemplateFPath,
            listMergeFileFPath,
            self.m_InputDataObj.OutputFileFPath,
            nValidDay
        )

    def DoMergeMembersToHalfMonthly(self, szMergeTemplateFPath, szSrcFDir, szMergeFileFPath, nValidDay):
        self.m_AppObj.Info(
            "\ntemplate path:{0}\nsrc dir:{1}\ndest file:{2}\nvalid day:{3}\n".format(
                szMergeTemplateFPath,
                szSrcFDir,
                szMergeFileFPath,
                nValidDay
            ))

        MrWorkbookObj = HalfMonthlyWorkbook(self.m_AppObj, szMergeTemplateFPath, szSrcFDir, nValidDay)
        MrWorkbookObj.Handle()
        MrWorkbookObj.Save(szMergeFileFPath)
        MrWorkbookObj.Close()

        self.m_AppObj.Info("End MergeHalfMonthlyReport")

    def DoMergeHalfMonthlyToMonthly(self, szOutTemplateFPath: str, listMergeFileFPath: list, szOutputFileFPath: str,
                                    nValidDay: int):
        logging.getLogger("myLog").info(
            "\ntemplate path:{0}\nhalf month report:{1}\noutput file:{2}\nvalid day:{3}".format(
                szOutTemplateFPath,
                "\n".join(listMergeFileFPath),
                szOutputFileFPath,
                nValidDay
            ))

        MrWorkbookObj = MonthlyWorkbook(self.m_AppObj, szOutTemplateFPath, listMergeFileFPath, nValidDay)
        MrWorkbookObj.Handle()
        MrWorkbookObj.Save(szOutputFileFPath)
        MrWorkbookObj.Close()

        self.m_AppObj.Info("End MergeMonthlyReport")
        pass


# HalfMonthlyReporWorkbook
class HalfMonthlyWorkbook:
    def __init__(self, AppObj, szTemplatePath, szSrcDir, nValidDay):
        self.m_AppObj = AppObj
        self.m_WorkbookObj = openpyxl.load_workbook(szTemplatePath)
        self.m_szSrcDir = szSrcDir
        self.m_ValidDay = nValidDay
        self.m_listMemberName = []

    def _CheckData(self):
        listSheetName = self.m_WorkbookObj.sheetnames
        self.m_AppObj.Info("{0}, {1}".format(ESheetName.eSummary, ESheetName.eTemplate))

        assert ESheetName.eSummary in listSheetName and ESheetName.eTemplate in listSheetName, \
            "tempalte excel missing sheet {0} or {1},  sheet name list: {2}".format(
                ESheetName.eSummary, ESheetName.eTemplate, ",".join(listSheetName))

        self.m_listMemberName = []
        SummarySheetObj = self.m_WorkbookObj[ESheetName.eSummary]
        nMaxColumn = SummarySheetObj.max_column
        for nColumnIndex in range(2, nMaxColumn + 1):
            szColumnChar = chr(96 + nColumnIndex)
            szCellPos1 = "{0}1".format(szColumnChar)
            szMemberName = SummarySheetObj[szCellPos1].value
            if szMemberName is not None:
                self.m_listMemberName.append(szMemberName)

        listMissingMemberExcel = []
        for szMemberName in self.m_listMemberName:
            szMemberFilePath = "{0}/{1}.xlsx".format(self.m_szSrcDir, szMemberName)
            if not os.path.exists(szMemberFilePath):
                listMissingMemberExcel.append(szMemberName)
                self.m_AppObj.Error("missing excel:{0}".format(szMemberFilePath))

        if len(listMissingMemberExcel) > 0:
            self.m_AppObj.Error("member missing excel:{0}".format(",".join(listMissingMemberExcel)))
            return False

        return True

    def Handle(self):
        self.m_AppObj.Debug("begin workbook handle")

        if not self._CheckData():
            raise my_exception.MyException("missing member excel")

        TemplateSheetObj = self.m_WorkbookObj[ESheetName.eTemplate]
        for szMemberName in self.m_listMemberName:
            szMemberFilePath = "{0}/{1}.xlsx".format(self.m_szSrcDir, szMemberName)
            MemberWorkbookObj = openpyxl.load_workbook(szMemberFilePath)
            szFirstSheetName = MemberWorkbookObj.sheetnames[0]
            SrcMemberSheetObj = MemberWorkbookObj[szFirstSheetName]

            MemberSheetObj = self.m_WorkbookObj.copy_worksheet(TemplateSheetObj)
            MemberSheetObj.title = szMemberName

            TmSheetObj = TmSheet(self.m_AppObj, MemberSheetObj, SrcMemberSheetObj, self.m_ValidDay)
            TmSheetObj.Handle()

        SheetObjInBook = self.m_WorkbookObj[ESheetName.eSummary]
        SummarySheetObj = SummarySheet(self.m_AppObj, SheetObjInBook, self.m_listMemberName)
        SummarySheetObj.Handle()

        self.m_AppObj.Debug("end workbook handle")

        if len(g_ErrorLog) > 0:
            print("报错信息:\n", "\n".join(g_ErrorLog))

    def Save(self, szPath):
        self.m_WorkbookObj.save(szPath)

    def Close(self):
        self.m_WorkbookObj.close()
        self.m_WorkbookObj = None


# TeamMemberSheet：每个成员表
class TmSheet:
    def __init__(self, AppObj, SheetObj, SrcSheetObj, nValidDay):
        self.m_AppObj = AppObj
        self.m_SheetObj = SheetObj
        self.m_SrcSheetObj = SrcSheetObj
        self.m_ValidDay = nValidDay

        self.m_nOldMaxRow = _GetMaxRow(self.m_SheetObj)
        self.m_nSrcMaxRow = _GetMaxRow(self.m_SrcSheetObj)
        self.m_nMaxRow = 0

    def Handle(self):
        self.m_AppObj.Debug("begin tmsheet handle:{0}".format(self.m_SheetObj.title))

        # 数据修改
        self._UpdateData()

        # 格式修改
        self._UpdateFormat()

        # 数据检查
        self._CheckDataValidate()

        self.m_AppObj.Debug("end tmsheet handle:{0}".format(self.m_SheetObj.title))

    def _UpdateData(self):
        self.m_AppObj.Debug("update data")

        nMaxRow = self.m_nSrcMaxRow
        # 时间
        nCurYear, nCurMonth = _GetCurYearMonth()
        self.m_SheetObj["A2"].value = "%4d-%02d" % (nCurYear, nCurMonth)

        # 工作内容
        # 天数
        for nRowIndex in range(2, nMaxRow + 1):
            szCellPos = "B{0}".format(nRowIndex)
            _CopyCell(self.m_SrcSheetObj[szCellPos], self.m_SheetObj[szCellPos])

            szCellPos = "C{0}".format(nRowIndex)
            _CopyCell(self.m_SrcSheetObj[szCellPos], self.m_SheetObj[szCellPos])

            if self.m_SheetObj[szCellPos].value is not None and self.m_SheetObj[szCellPos] != "":
                self.m_nMaxRow = nRowIndex

        self.m_AppObj.Info("max row:{0}".format(self.m_nMaxRow))

        # 重要程度
        # 代码质量
        # 实现难度
        # 交付效率
        listCharacter = ["D", "E", "F", "G"]
        for nRowIndex in range(3, self.m_nMaxRow + 1):
            for szChar in listCharacter:
                szDefaultPos = "{0}2".format(szChar)
                szCellPos = "{0}{1}".format(szChar, nRowIndex)
                _CopyCell(self.m_SheetObj[szDefaultPos], self.m_SheetObj[szCellPos])

        # 评分：公式
        szScoreFormat = "=C{0} * IF(D{0} =\"核心\",1.3,IF(D{0}=\"基本\",1.1,IF(D{0}=\"次要\",0.9,IF(D{0}=\"周边\",0.7," \
                        "IF(D{0}=\"改bug\",0.5,IF(D{0}=\"自学\",0.2,IF(D{0}=\"无关\",0,0))))))) \
            * IF(E{0} =\"超水准\",1.3,IF(E{0}=\"基本达标\",1,IF(E{0}=\"少量问题\",0.8,IF(E{0}=\"引发事故\",0.6,IF(E{0}=\"\",0))))) \
            * IF(F{0} =\"噩梦\",1.5,IF(F{0}=\"困难\",1.3,IF(F{0}=\"普通\",1,IF(F{0}=\"简单\",0.8,IF(F{0}=\"小白\",0.6,0)))) \
            * IF(G{0} =\"超前\",1.2,IF(G{0}=\"按时\",1,IF(G{0}=\"稍晚\",0.8,IF(G{0}=\"延期\",0.6,IF(G{0}=\"中止\",0.5,0))))))"

        for nRowIndex in range(2, self.m_nMaxRow + 1):
            szCellPos = "H{0}".format(nRowIndex)
            self.m_SheetObj[szCellPos].value = szScoreFormat.format(nRowIndex)

        # 微调

        # 总计
        self.m_SheetObj["J2"] = "=SUM(H2:H{0}, I2)".format(nMaxRow)

    def Merge(self):
        self.m_AppObj.Debug("begin tmsheet merge:{0}".format(self.m_SheetObj.title))

        # 数据修改
        self._MergeData()

        # 格式修改
        self._UpdateFormat()

        # 数据检查
        self._CheckDataValidate()

        self.m_AppObj.Debug("end tmsheet merge:{0}".format(self.m_SheetObj.title))

    def _MergeData(self):
        self.m_AppObj.Debug("update data")

        nSrcMaxRow = self.m_nSrcMaxRow

        # 时间
        nCurYear, nCurMonth = _GetCurYearMonth()
        self.m_SheetObj["A2"].value = "%4d-%02d" % (nCurYear, nCurMonth)

        # 工作内容
        # 天数
        # 重要程度
        # 代码质量
        # 实现难度
        # 交付效率
        listCharacter = ["B", "C", "D", "E", "F", "G"]
        for szChar in listCharacter:
            for nRowIndex in range(2, nSrcMaxRow + 1):
                szSrcCellPos = "{0}{1}".format(szChar, nRowIndex)
                szCellPos = "{0}{1}".format(szChar, self.m_nOldMaxRow + nRowIndex - 1)

                _CopyCell(self.m_SrcSheetObj[szSrcCellPos], self.m_SheetObj[szCellPos])

                if self.m_SheetObj[szCellPos].value is not None:
                    self.m_nMaxRow = self.m_nOldMaxRow + nRowIndex - 1

        self.m_AppObj.Info("max row:{0}".format(self.m_nMaxRow))

        # 评分：公式
        szScoreFormat = "=C{0} * IF(D{0} =\"核心\",1.3,IF(D{0}=\"基本\",1.1,IF(D{0}=\"次要\",0.9,IF(D{0}=\"周边\",0.7," \
                        "IF(D{0}=\"改bug\",0.5,IF(D{0}=\"自学\",0.2,IF(D{0}=\"无关\",0,0))))))) \
            * IF(E{0} =\"超水准\",1.3,IF(E{0}=\"基本达标\",1,IF(E{0}=\"少量问题\",0.8,IF(E{0}=\"引发事故\",0.6,IF(E{0}=\"\",0))))) \
            * IF(F{0} =\"噩梦\",1.5,IF(F{0}=\"困难\",1.3,IF(F{0}=\"普通\",1,IF(F{0}=\"简单\",0.8,IF(F{0}=\"小白\",0.6,0)))) \
            * IF(G{0} =\"超前\",1.2,IF(G{0}=\"按时\",1,IF(G{0}=\"稍晚\",0.8,IF(G{0}=\"延期\",0.6,IF(G{0}=\"中止\",0.5,0))))))"

        for nRowIndex in range(2, self.m_nMaxRow + 1):
            szCellPos = "H{0}".format(nRowIndex)
            self.m_SheetObj[szCellPos].value = szScoreFormat.format(nRowIndex)

        # 微调
        szWeiTiaoPos = "I2"
        if not _IsCellValid(self.m_SheetObj[szWeiTiaoPos]) or not isinstance(self.m_SheetObj[szWeiTiaoPos].value, int):
            _CopyCell(self.m_SrcSheetObj[szWeiTiaoPos], self.m_SheetObj[szWeiTiaoPos])
        else:
            self.m_SheetObj[szWeiTiaoPos].value += self.m_SrcSheetObj[szWeiTiaoPos].value

        # 总计
        self.m_SheetObj["J2"] = "=SUM(H2:H{0}, I2)".format(self.m_nMaxRow)

    def _UpdateFormat(self):
        self.m_AppObj.Debug("update format")

        nCurYear, nCurMonth = _GetCurYearMonth()
        self.m_SheetObj["A2"].value = "%4d-%02d" % (nCurYear, nCurMonth)

        nMaxRow = self.m_nMaxRow
        nOldMaxRow = self.m_nOldMaxRow
        self.m_AppObj.Debug("max row:{0}, old max row:{1}".format(nMaxRow, nOldMaxRow))

        # 时间合并单元格
        # 微调合并单元格
        # 总计合并单元格
        listChar = ['A', 'I', 'J']
        for szChar in listChar:
            szOldCellRange = "{0}2:{0}{1}".format(szChar, nOldMaxRow)
            szCellRange = "{0}2:{0}{1}".format(szChar, nMaxRow)
            try:
                self.m_SheetObj.unmerge_cells(szOldCellRange)
            except ValueError:
                # logging.getLogger("myLog").debug("unmerge_cells failed:%s", szCellRange)
                pass

            self.m_SheetObj.merge_cells(szCellRange)
            logging.getLogger("myLog").debug("cell range:%s", szCellRange)

        # 更新数据验证
        dictDataValidation = {
            "D": '"核心,基本,次要,周边,改bug,自学,无关"',
            "E": '"超水准,基本达标,少量问题,引发事故"',
            "F": '"噩梦,困难,普通,简单,小白"',
            "G": '"超前,按时,稍晚,延期,中止"'
        }
        for szColName, szData in dictDataValidation.items():
            DataValidationObj = DataValidation(type="list", formula1=szData, allow_blank=True)
            DataValidationObj.add('{0}1:{0}1048576'.format(szColName))
            self.m_SheetObj.add_data_validation(DataValidationObj)

    def _CheckDataValidate(self):
        self.m_AppObj.Debug("update data validate")

        # 天数检查
        nSum = 0
        nMaxRow = self.m_nMaxRow
        for nRowIndex in range(2, nMaxRow + 1):
            try:
                nSum += int(self.m_SheetObj["C{0}".format(nRowIndex)].value)
            except TypeError:
                szError = "天数类型错误:{0}, {1}".format(self.m_SheetObj.title, "C{0}".format(nRowIndex))
                self.m_AppObj.Error(szError)
                g_ErrorLog.append(szError)

        if nSum > self.m_ValidDay:
            szError = "天数总和异常:{0}, {1}".format(self.m_SheetObj.title, nSum)
            self.m_AppObj.Error(szError)
            g_ErrorLog.append(szError)


# 汇总表
class SummarySheet:
    def __init__(self, AppObj, SheetObj, listMemberName):
        self.m_AppObj = AppObj
        self.m_SheetObj = SheetObj
        self.m_listMemberName = listMemberName

    def Handle(self):
        self.m_AppObj.Debug("Begin Handle")

        nMaxColumn = len(self.m_listMemberName) + 1
        for nColumnIndex in range(2, nMaxColumn + 1):
            szColumnChar = chr(96 + nColumnIndex)
            szCellPos1 = "{0}1".format(szColumnChar)
            szCellPos2 = "{0}2".format(szColumnChar)
            self.m_SheetObj[szCellPos2].value = "={0}!J2".format(self.m_SheetObj[szCellPos1].value)


# MonthlyReporWorkbook
class MonthlyWorkbook:
    def __init__(self, AppObj, szTemplatePath, listMergeFileFPath, nValidDay):
        self.m_AppObj = AppObj
        self.m_WorkbookObj = openpyxl.load_workbook(szTemplatePath)
        self.m_szTemplatePath = szTemplatePath
        self.m_listMergeFileFPath = listMergeFileFPath
        self.m_nValidDay = nValidDay

        self.m_listMemberName = []

    def _CheckData(self):
        listSheetName = self.m_WorkbookObj.sheetnames
        self.m_AppObj.Info("{0}, {1}".format(ESheetName.eSummary, ESheetName.eTemplate))

        assert ESheetName.eSummary in listSheetName and ESheetName.eTemplate in listSheetName, \
            "tempalte excel missing sheet {0} or {1},  sheet name list: {2}".format(
                ESheetName.eSummary, ESheetName.eTemplate, ",".join(listSheetName))

        self.m_listMemberName = []
        SummarySheetObj = self.m_WorkbookObj[ESheetName.eSummary]
        nMaxColumn = SummarySheetObj.max_column
        for nColumnIndex in range(2, nMaxColumn + 1):
            szColumnChar = chr(96 + nColumnIndex)
            szCellPos1 = "{0}1".format(szColumnChar)
            szMemberName = SummarySheetObj[szCellPos1].value
            if szMemberName is not None:
                self.m_listMemberName.append(szMemberName)

        listMissingMergeFileFPath = []
        for szMergeFileFPath in self.m_listMergeFileFPath:
            if not os.path.exists(szMergeFileFPath):
                listMissingMergeFileFPath.append(szMergeFileFPath)
                self.m_AppObj.Error("missing excel:{0}".format(szMergeFileFPath))

        if len(listMissingMergeFileFPath) > 0:
            self.m_AppObj.Error("merge file missing excel:{0}".format(",".join(listMissingMergeFileFPath)))
            return False

        return True

    def Handle(self):
        self.m_AppObj.Debug("begin workbook handle")

        if not self._CheckData():
            raise my_exception.MyException("missing member excel")

        self._InitAllTmSheet()

        for szMergeFileFPath in self.m_listMergeFileFPath:
            self._MergeFileTmSheet(szMergeFileFPath)

        SheetObjInBook = self.m_WorkbookObj[ESheetName.eSummary]
        SummarySheetObj = SummarySheet(self.m_AppObj, SheetObjInBook, self.m_listMemberName)
        SummarySheetObj.Handle()

        self.m_AppObj.Debug("end workbook handle")

        if len(g_ErrorLog) > 0:
            print("报错信息:\n", "\n".join(g_ErrorLog))

    def _InitAllTmSheet(self):
        TemplateSheetObj = self.m_WorkbookObj[ESheetName.eTemplate]
        for szMemberName in self.m_listMemberName:
            MemberSheetObj = self.m_WorkbookObj.copy_worksheet(TemplateSheetObj)
            MemberSheetObj.title = szMemberName
            logging.getLogger("myLog").debug("copy sheet:%s", szMemberName)

    def _MergeFileTmSheet(self, szMergeFileFPath):
        logging.getLogger("myLog").info("Merge file:%s", szMergeFileFPath)

        MergeFileWorkbookObj = openpyxl.load_workbook(szMergeFileFPath)

        for szMemberName in self.m_listMemberName:
            MemberSheetObj = self.m_WorkbookObj[szMemberName]
            SrcMemberSheetObj = MergeFileWorkbookObj[szMemberName]

            MonthlyTmSheetObj = TmSheet(self.m_AppObj, MemberSheetObj, SrcMemberSheetObj, self.m_nValidDay)
            MonthlyTmSheetObj.Merge()

    def Save(self, szPath):
        self.m_WorkbookObj.save(szPath)

    def Close(self):
        self.m_WorkbookObj.close()
        self.m_WorkbookObj = None

# 单元格填充自动拓展算法：
# 1、匹配特殊序列
# 2、增长规律
